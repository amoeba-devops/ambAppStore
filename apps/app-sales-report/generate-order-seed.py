#!/usr/bin/env python3
"""
Generate seed SQL for raw orders from Shopee Excel export.

Usage:
    python3 apps/app-sales-report/generate-order-seed.py

Input:  reference/appSalesReport/Order.all.20260301_20260331.xlsx
Output: apps/app-sales-report/seed-orders-202603.sql

Reads existing SKU master data from seed-product-master.sql to build
wms_code → sku_id mapping for automatic SKU matching.
"""
import openpyxl
import uuid
import re
import os
from collections import OrderedDict

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
EXCEL_PATH = 'reference/appSalesReport/Order.all.20260301_20260331-1.xlsx'
OUTPUT_PATH = 'apps/app-sales-report/seed-orders-202603.sql'
SEED_SQL_PATH = 'apps/app-sales-report/seed-product-master.sql'
ENT_ID = 'acce6566-8a00-4071-b52b-082b69832510'
CHN_CODE = 'SHOPEE'
BATCH_ID = 'import-20260301-20260331'
CHUNK_SIZE = 500  # rows per INSERT batch

# ---------------------------------------------------------------------------
# Status mapping: Vietnamese → Normalized ENUM
# ---------------------------------------------------------------------------
STATUS_MAP = {
    'Hoàn thành': 'COMPLETED',
    'Đã hủy': 'CANCELLED',
    'Đã giao': 'DELIVERED',
    'Đang giao': 'SHIPPING',
    'Đã nhận được hàng': 'RECEIVED',
}

def normalize_status(raw_status):
    if not raw_status:
        return 'UNKNOWN', ''
    s = raw_status.strip()
    for vi, en in STATUS_MAP.items():
        if s == vi:
            return en, s
    if 'Người mua xác nhận đã nhận được hàng' in s:
        return 'DELIVERED_REFUNDABLE', s
    return 'UNKNOWN', s

# ---------------------------------------------------------------------------
# Return status mapping
# ---------------------------------------------------------------------------
RETURN_STATUS_MAP = {
    'Đã Chấp Thuận Yêu Cầu': 'APPROVED',
    'Hoàn tất trả hàng': 'RETURNED',
    'Đã giải quyết khiếu nại': 'RESOLVED',
    'Yêu cầu chờ xử lý': 'PENDING',
}

def normalize_return_status(raw):
    if not raw or not str(raw).strip():
        return None
    s = str(raw).strip()
    return RETURN_STATUS_MAP.get(s, s[:50])

# ---------------------------------------------------------------------------
# SQL helpers
# ---------------------------------------------------------------------------
def esc(s):
    """Escape string for SQL single-quoted value."""
    if s is None:
        return None
    v = str(s).strip().replace('\\', '\\\\').replace("'", "\\'").replace('\n', ' ').replace('\r', '').replace('\t', ' ')
    return v if v else None

def sql_str(v):
    """Return SQL-safe quoted string or NULL."""
    e = esc(v)
    return f"'{e}'" if e is not None else 'NULL'

def sql_decimal(v):
    """Return decimal or NULL."""
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if not s:
        return 'NULL'
    try:
        return f"'{float(s):.2f}'"
    except (ValueError, TypeError):
        return 'NULL'

def sql_decimal3(v):
    """Return decimal(8,3) or NULL."""
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if not s:
        return 'NULL'
    try:
        return f"'{float(s):.3f}'"
    except (ValueError, TypeError):
        return 'NULL'

def sql_int(v):
    if v is None:
        return 'NULL'
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return 'NULL'

def sql_datetime(v):
    """Convert datetime string to SQL DATETIME or NULL."""
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if not s:
        return 'NULL'
    # Handle 'YYYY-MM-DD HH:MM' (no seconds)
    if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}$', s):
        return f"'{s}:00'"
    # Handle 'YYYY-MM-DD HH:MM:SS'
    if re.match(r'^\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2}', s):
        return f"'{s[:19]}'"
    return 'NULL'

def sql_bool(v):
    """Map 'Y'/'N' or truthy to 0/1."""
    if v is None:
        return '0'
    s = str(v).strip().upper()
    return '1' if s in ('Y', 'YES', 'TRUE', '1') else '0'

# ---------------------------------------------------------------------------
# Load SKU mapping from existing seed SQL
# ---------------------------------------------------------------------------
def load_sku_mapping(seed_path):
    """Parse seed-product-master.sql to build wms_code → sku_id mapping."""
    mapping = {}
    if not os.path.exists(seed_path):
        print(f"WARNING: Seed file not found: {seed_path}")
        return mapping

    with open(seed_path, 'r', encoding='utf-8') as f:
        for line in f:
            if 'INSERT INTO drd_sku_masters' in line:
                # Extract sku_id and sku_wms_code from VALUES(...)
                m = re.search(r"VALUES\s*\('([^']+)',\s*'[^']+',\s*'[^']+',\s*'([^']+)'", line)
                if m:
                    sku_id = m.group(1)
                    wms_code = m.group(2)
                    mapping[wms_code] = sku_id
    return mapping

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=False)
    ws = wb['orders']

    # Load SKU mapping
    sku_map = load_sku_mapping(SEED_SQL_PATH)
    print(f"Loaded {len(sku_map)} SKU mappings from seed SQL")

    # -----------------------------------------------------------------------
    # Pass 1: Group rows by order ID
    # -----------------------------------------------------------------------
    orders = OrderedDict()  # order_id → { 'first_row': [...], 'items': [[...], ...] }
    row_count = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        order_id = row[0]
        if not order_id:
            continue
        order_id = str(order_id).strip()
        row_count += 1

        if order_id not in orders:
            orders[order_id] = {'first_row': list(row), 'items': []}
        orders[order_id]['items'].append(list(row))

    print(f"Total rows: {row_count}, Unique orders: {len(orders)}")

    # -----------------------------------------------------------------------
    # Pass 2: Generate SQL
    # -----------------------------------------------------------------------
    lines = []
    lines.append("-- ============================================================")
    lines.append("-- Seed Data: Raw Orders (Shopee Vietnam)")
    lines.append(f"-- Source: {os.path.basename(EXCEL_PATH)}")
    lines.append(f"-- Entity: {ENT_ID}")
    lines.append(f"-- Channel: {CHN_CODE}")
    lines.append(f"-- Batch: {BATCH_ID}")
    lines.append(f"-- Total Orders: {len(orders)}, Total Items: {row_count}")
    lines.append("-- Generated by generate-order-seed.py")
    lines.append("-- ============================================================")
    lines.append("")
    lines.append("SET NAMES utf8mb4;")
    lines.append("USE db_app_sales;")
    lines.append("")

    # DDL — Create tables if not exist
    lines.append("-- ---- DDL: Create tables ----")
    lines.append("""
CREATE TABLE IF NOT EXISTS drd_raw_orders (
  ord_id                        CHAR(36)        NOT NULL,
  ent_id                        CHAR(36)        NOT NULL,
  chn_code                      VARCHAR(20)     NOT NULL,
  ord_channel_order_id          VARCHAR(30)     NOT NULL,
  ord_package_id                VARCHAR(30)     NULL,
  ord_order_date                DATETIME        NOT NULL,
  ord_status                    VARCHAR(20)     NOT NULL,
  ord_status_raw                VARCHAR(500)    NULL,
  ord_cancel_reason             VARCHAR(500)    NULL,
  ord_tracking_no               VARCHAR(50)     NULL,
  ord_carrier                   VARCHAR(100)    NULL,
  ord_delivery_method           VARCHAR(50)     NULL,
  ord_order_type                VARCHAR(50)     NULL,
  ord_est_delivery_date         DATETIME        NULL,
  ord_ship_date                 DATETIME        NULL,
  ord_delivery_time             DATETIME        NULL,
  ord_total_weight_kg           DECIMAL(8,3)    NULL,
  ord_total_vnd                 DECIMAL(15,2)   NULL,
  ord_shop_voucher              VARCHAR(100)    NULL,
  ord_coin_cashback             DECIMAL(15,2)   NULL,
  ord_shopee_voucher            VARCHAR(100)    NULL,
  ord_promo_combo               VARCHAR(200)    NULL,
  ord_shopee_combo_discount     DECIMAL(15,2)   NULL,
  ord_shop_combo_discount       DECIMAL(15,2)   NULL,
  ord_shopee_coin_rebate        DECIMAL(15,2)   NULL,
  ord_card_discount             DECIMAL(15,2)   NULL,
  ord_trade_in_discount         DECIMAL(15,2)   NULL,
  ord_trade_in_bonus            DECIMAL(15,2)   NULL,
  ord_seller_trade_in_bonus     DECIMAL(15,2)   NULL,
  ord_shipping_fee_est          DECIMAL(15,2)   NULL,
  ord_buyer_shipping_fee        DECIMAL(15,2)   NULL,
  ord_shopee_shipping_subsidy   DECIMAL(15,2)   NULL,
  ord_return_shipping_fee       DECIMAL(15,2)   NULL,
  ord_total_buyer_payment       DECIMAL(15,2)   NULL,
  ord_completed_at              DATETIME        NULL,
  ord_paid_at                   DATETIME        NULL,
  ord_payment_method            VARCHAR(100)    NULL,
  ord_commission_fee            DECIMAL(15,2)   NULL,
  ord_service_fee               DECIMAL(15,2)   NULL,
  ord_payment_fee               DECIMAL(15,2)   NULL,
  ord_deposit                   DECIMAL(15,2)   NULL,
  ord_province                  VARCHAR(100)    NULL,
  ord_district                  VARCHAR(100)    NULL,
  ord_country                   VARCHAR(50)     NULL,
  ord_import_batch_id           VARCHAR(50)     NULL,
  ord_created_at                DATETIME        NOT NULL DEFAULT CURRENT_TIMESTAMP,
  ord_updated_at                DATETIME        NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  CONSTRAINT pk_drd_raw_orders PRIMARY KEY (ord_id),
  CONSTRAINT uq_drd_raw_orders_channel_order UNIQUE (ent_id, chn_code, ord_channel_order_id),
  INDEX idx_drd_raw_orders_date (ent_id, ord_order_date),
  INDEX idx_drd_raw_orders_status (ent_id, ord_status),
  INDEX idx_drd_raw_orders_batch (ord_import_batch_id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
""")

    lines.append("""
CREATE TABLE IF NOT EXISTS drd_raw_order_items (
  oli_id                    CHAR(36)        NOT NULL,
  ent_id                    CHAR(36)        NOT NULL,
  ord_id                    CHAR(36)        NOT NULL,
  sku_id                    CHAR(36)        NULL,
  oli_product_sku           VARCHAR(50)     NULL,
  oli_product_name          VARCHAR(500)    NULL,
  oli_variant_sku           VARCHAR(30)     NULL,
  oli_variant_name          VARCHAR(200)    NULL,
  oli_is_bestseller         TINYINT(1)      NOT NULL DEFAULT 0,
  oli_weight_kg             DECIMAL(8,3)    NULL,
  oli_original_price        DECIMAL(15,2)   NULL,
  oli_seller_discount       DECIMAL(15,2)   NULL,
  oli_shopee_discount       DECIMAL(15,2)   NULL,
  oli_total_seller_subsidy  DECIMAL(15,2)   NULL,
  oli_deal_price            DECIMAL(15,2)   NULL,
  oli_quantity              INT             NOT NULL DEFAULT 1,
  oli_return_quantity       INT             NULL DEFAULT 0,
  oli_buyer_paid            DECIMAL(15,2)   NULL,
  oli_return_status         VARCHAR(50)     NULL,
  oli_sku_match_status      VARCHAR(10)     NOT NULL DEFAULT 'UNMATCHED',
  oli_created_at            DATETIME        NOT NULL DEFAULT CURRENT_TIMESTAMP,
  oli_updated_at            DATETIME        NOT NULL DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
  CONSTRAINT pk_drd_raw_order_items PRIMARY KEY (oli_id),
  CONSTRAINT fk_drd_raw_order_items_order FOREIGN KEY (ord_id) REFERENCES drd_raw_orders (ord_id),
  CONSTRAINT fk_drd_raw_order_items_sku FOREIGN KEY (sku_id) REFERENCES drd_sku_masters (sku_id),
  INDEX idx_drd_raw_order_items_ord (ent_id, ord_id),
  INDEX idx_drd_raw_order_items_sku (ent_id, oli_variant_sku),
  INDEX idx_drd_raw_order_items_sku_id (sku_id)
) ENGINE=InnoDB DEFAULT CHARSET=utf8mb4 COLLATE=utf8mb4_unicode_ci;
""")

    # -----------------------------------------------------------------------
    # Generate INSERT statements — Orders first, then Items
    # -----------------------------------------------------------------------
    lines.append("-- ---- Orders ----")

    order_lines = []
    item_lines = []
    match_stats = {'MATCHED': 0, 'UNMATCHED': 0, 'COMBO': 0}
    unmatched_skus = set()

    for order_id, data in orders.items():
        r = data['first_row']
        ord_uuid = str(uuid.uuid4())

        status, status_raw = normalize_status(r[3])
        order_date = sql_datetime(r[2])

        order_lines.append(
            f"INSERT INTO drd_raw_orders "
            f"(ord_id, ent_id, chn_code, ord_channel_order_id, ord_package_id, "
            f"ord_order_date, ord_status, ord_status_raw, ord_cancel_reason, "
            f"ord_tracking_no, ord_carrier, ord_delivery_method, ord_order_type, "
            f"ord_est_delivery_date, ord_ship_date, ord_delivery_time, "
            f"ord_total_weight_kg, ord_total_vnd, "
            f"ord_shop_voucher, ord_coin_cashback, ord_shopee_voucher, "
            f"ord_promo_combo, ord_shopee_combo_discount, ord_shop_combo_discount, "
            f"ord_shopee_coin_rebate, ord_card_discount, "
            f"ord_trade_in_discount, ord_trade_in_bonus, ord_seller_trade_in_bonus, "
            f"ord_shipping_fee_est, ord_buyer_shipping_fee, ord_shopee_shipping_subsidy, "
            f"ord_return_shipping_fee, ord_total_buyer_payment, "
            f"ord_completed_at, ord_paid_at, ord_payment_method, "
            f"ord_commission_fee, ord_service_fee, ord_payment_fee, ord_deposit, "
            f"ord_province, ord_district, ord_country, "
            f"ord_import_batch_id) "
            f"VALUES ("
            f"'{ord_uuid}', '{ENT_ID}', '{CHN_CODE}', {sql_str(order_id)}, {sql_str(r[1])}, "
            f"{order_date}, '{status}', {sql_str(status_raw)}, {sql_str(r[5])}, "
            f"{sql_str(r[7])}, {sql_str(r[8])}, {sql_str(r[9])}, {sql_str(r[10])}, "
            f"{sql_datetime(r[11])}, {sql_datetime(r[12])}, {sql_datetime(r[13])}, "
            f"{sql_decimal3(r[18])}, {sql_decimal(r[29])}, "
            f"{sql_str(r[30])}, {sql_decimal(r[31])}, {sql_str(r[32])}, "
            f"{sql_str(r[33])}, {sql_decimal(r[34])}, {sql_decimal(r[35])}, "
            f"{sql_decimal(r[36])}, {sql_decimal(r[37])}, "
            f"{sql_decimal(r[38])}, {sql_decimal(r[39])}, {sql_decimal(r[41])}, "
            f"{sql_decimal(r[40])}, {sql_decimal(r[42])}, {sql_decimal(r[43])}, "
            f"{sql_decimal(r[44])}, {sql_decimal(r[45])}, "
            f"{sql_datetime(r[46])}, {sql_datetime(r[47])}, {sql_str(r[48])}, "
            f"{sql_decimal(r[49])}, {sql_decimal(r[50])}, {sql_decimal(r[51])}, {sql_decimal(r[52])}, "
            f"{sql_str(r[56])}, {sql_str(r[57])}, {sql_str(r[60])}, "
            f"'{BATCH_ID}');"
        )

        # Items
        for item_row in data['items']:
            oli_uuid = str(uuid.uuid4())
            variant_sku = esc(item_row[19])  # col 20: SKU phân loại hàng

            # SKU matching
            sku_id = 'NULL'
            match_status = 'UNMATCHED'
            if variant_sku:
                if variant_sku.startswith('Combo') or '_GIFT_' in variant_sku.upper():
                    match_status = 'COMBO'
                    match_stats['COMBO'] += 1
                elif variant_sku in sku_map:
                    sku_id = f"'{sku_map[variant_sku]}'"
                    match_status = 'MATCHED'
                    match_stats['MATCHED'] += 1
                else:
                    match_stats['UNMATCHED'] += 1
                    unmatched_skus.add(variant_sku)
            else:
                match_stats['UNMATCHED'] += 1

            return_status = normalize_return_status(item_row[14])

            item_lines.append(
                f"INSERT INTO drd_raw_order_items "
                f"(oli_id, ent_id, ord_id, sku_id, "
                f"oli_product_sku, oli_product_name, oli_variant_sku, oli_variant_name, "
                f"oli_is_bestseller, oli_weight_kg, "
                f"oli_original_price, oli_seller_discount, oli_shopee_discount, "
                f"oli_total_seller_subsidy, oli_deal_price, "
                f"oli_quantity, oli_return_quantity, oli_buyer_paid, "
                f"oli_return_status, oli_sku_match_status) "
                f"VALUES ("
                f"'{oli_uuid}', '{ENT_ID}', '{ord_uuid}', {sku_id}, "
                f"{sql_str(item_row[15])}, {sql_str(item_row[16])}, {sql_str(item_row[19])}, {sql_str(item_row[20])}, "
                f"{sql_bool(item_row[4])}, {sql_decimal3(item_row[17])}, "
                f"{sql_decimal(item_row[21])}, {sql_decimal(item_row[22])}, {sql_decimal(item_row[23])}, "
                f"{sql_decimal(item_row[24])}, {sql_decimal(item_row[25])}, "
                f"{sql_int(item_row[26])}, {sql_int(item_row[27])}, {sql_decimal(item_row[28])}, "
                f"{sql_str(return_status)}, '{match_status}');"
            )

    # Write orders first, then items (FK constraint order)
    lines.extend(order_lines)
    lines.append("")
    lines.append("-- ---- Order Items ----")
    lines.extend(item_lines)

    lines.append("")

    # Write output
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    # Summary
    total_items = sum(match_stats.values())
    print(f"\n{'='*60}")
    print(f"Generated: {OUTPUT_PATH}")
    print(f"  Orders:  {len(orders)}")
    print(f"  Items:   {total_items}")
    print(f"  MATCHED:   {match_stats['MATCHED']} ({match_stats['MATCHED']*100//max(total_items,1)}%)")
    print(f"  UNMATCHED: {match_stats['UNMATCHED']} ({match_stats['UNMATCHED']*100//max(total_items,1)}%)")
    print(f"  COMBO:     {match_stats['COMBO']} ({match_stats['COMBO']*100//max(total_items,1)}%)")
    if unmatched_skus:
        print(f"\n  Unmatched SKUs ({len(unmatched_skus)}):")
        for s in sorted(unmatched_skus):
            print(f"    - {s}")
    print(f"{'='*60}")

if __name__ == '__main__':
    main()
