#!/usr/bin/env python3
"""
Generate seed SQL for raw orders from TikTok Shop Excel export.

Usage:
    python3 apps/app-sales-report/generate-tiktok-order-seed.py

Input:  reference/appSalesReport/tiktok-sales_report.xlsx
Output: apps/app-sales-report/seed-tiktok-orders-202603.sql
"""
import openpyxl
import uuid
import re
import os
from collections import OrderedDict

# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------
EXCEL_PATH = 'reference/appSalesReport/tiktok-sales_report.xlsx'
OUTPUT_PATH = 'apps/app-sales-report/seed-tiktok-orders-202603.sql'
SEED_SQL_PATH = 'apps/app-sales-report/seed-product-master.sql'
ENT_ID = 'acce6566-8a00-4071-b52b-082b69832510'
CHN_CODE = 'TIKTOK'
BATCH_ID = 'tiktok-import-202603'

# ---------------------------------------------------------------------------
# TikTok column indices (0-based)
# ---------------------------------------------------------------------------
COL_ORDER_ID = 0
COL_STATUS = 1
COL_SUBSTATUS = 2
COL_CANCEL_RETURN_TYPE = 3
COL_ORDER_TYPE = 4        # Normal or Pre-order
COL_SKU_ID = 5            # Platform SKU ID (skip)
COL_SELLER_SKU = 6        # Seller SKU → SKU matching key
COL_PRODUCT_NAME = 7
COL_VARIATION = 8
COL_QUANTITY = 9
COL_RETURN_QTY = 10
COL_UNIT_PRICE = 11       # SKU Unit Original Price
COL_SUBTOTAL_BEFORE = 12  # Before discount (computed, skip)
COL_PLATFORM_DISCOUNT = 13
COL_SELLER_DISCOUNT = 14
COL_SUBTOTAL_AFTER = 15   # After discount → deal price
COL_SHIP_FEE_AFTER = 16   # Shipping Fee After Discount
COL_SHIP_FEE_ORIG = 17    # Original Shipping Fee
COL_SHIP_FEE_SELLER = 18  # Shipping Fee Seller Discount
COL_SHIP_FEE_PLATFORM = 19
COL_PAYMENT_PLATFORM_DISC = 20
COL_TAXES = 21
COL_ORDER_AMOUNT = 22     # ord_total_vnd
COL_REFUND_AMOUNT = 23
COL_CREATED_TIME = 24     # Order Date
COL_PAID_TIME = 25
COL_RTS_TIME = 26         # Ready to Ship
COL_SHIPPED_TIME = 27
COL_DELIVERED_TIME = 28
COL_CANCELLED_TIME = 29
COL_CANCEL_BY = 30
COL_CANCEL_REASON = 31
COL_FULFILLMENT = 32      # Fulfillment Type
COL_WAREHOUSE = 33
COL_TRACKING_ID = 34
COL_DELIVERY_OPTION = 35
COL_CARRIER = 36          # Shipping Provider Name
# 37-40: PII (skip)
COL_COUNTRY = 41
COL_PROVINCE = 42
COL_DISTRICT = 43
# 44-46: Address PII (skip)
COL_PAYMENT_METHOD = 47
COL_WEIGHT_KG = 48        # Item-level weight
COL_CATEGORY = 49
COL_PACKAGE_ID = 50

# ---------------------------------------------------------------------------
# Status mapping: Vietnamese → Normalized ENUM
# ---------------------------------------------------------------------------
STATUS_MAP = {
    'Đã hoàn tất': 'COMPLETED',
    'Đã hủy': 'CANCELLED',
    'Đã vận chuyển': 'SHIPPING',
    'Đã giao': 'DELIVERED',
    'Đang vận chuyển': 'SHIPPING',
}

def normalize_status(raw_status):
    if not raw_status:
        return 'UNKNOWN', ''
    s = str(raw_status).strip()
    for vi, en in STATUS_MAP.items():
        if s == vi:
            return en, s
    return 'UNKNOWN', s

# ---------------------------------------------------------------------------
# SQL helpers
# ---------------------------------------------------------------------------
def esc(s):
    if s is None:
        return None
    v = str(s).strip().replace('\\', '\\\\').replace("'", "\\'").replace('\n', ' ').replace('\r', '').replace('\t', ' ')
    return v if v else None

def sql_str(v):
    e = esc(v)
    return f"'{e}'" if e is not None else 'NULL'

def sql_decimal(v):
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if not s:
        return 'NULL'
    try:
        return f"'{float(s):.2f}'"
    except (ValueError, TypeError):
        return 'NULL'

def sql_decimal3(v):
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if not s:
        return 'NULL'
    try:
        return f"'{float(s):.3f}'"
    except (ValueError, TypeError):
        return 'NULL'

def sql_int(v):
    if v is None:
        return 'NULL'
    try:
        return str(int(float(str(v).strip())))
    except (ValueError, TypeError):
        return 'NULL'

def sql_datetime_tiktok(v):
    """Convert TikTok datetime format DD/MM/YYYY HH:mm:ss to SQL DATETIME."""
    if v is None:
        return 'NULL'
    s = str(v).strip()
    if not s:
        return 'NULL'
    # TikTok format: DD/MM/YYYY HH:mm:ss
    m = re.match(r'^(\d{2})/(\d{2})/(\d{4})\s+(\d{2}:\d{2}:\d{2})$', s)
    if m:
        day, month, year, time = m.groups()
        return f"'{year}-{month}-{day} {time}'"
    # Fallback: ISO format
    if re.match(r'^\d{4}-\d{2}-\d{2}', s):
        return f"'{s[:19]}'"
    return 'NULL'

# ---------------------------------------------------------------------------
# Load SKU mapping from existing seed SQL
# ---------------------------------------------------------------------------
def load_sku_mapping(seed_path):
    mapping = {}
    if not os.path.exists(seed_path):
        print(f"WARNING: Seed file not found: {seed_path}")
        return mapping
    with open(seed_path, 'r', encoding='utf-8') as f:
        for line in f:
            if 'INSERT INTO drd_sku_masters' in line:
                m = re.search(r"VALUES\s*\('([^']+)',\s*'[^']+',\s*'[^']+',\s*'([^']+)'", line)
                if m:
                    sku_id = m.group(1)
                    wms_code = m.group(2)
                    mapping[wms_code] = sku_id
    return mapping

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    print(f"Loading Excel: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    sku_map = load_sku_mapping(SEED_SQL_PATH)
    print(f"Loaded {len(sku_map)} SKU mappings from seed SQL")

    # Group rows by order ID (skip row 2 = description row)
    orders = OrderedDict()
    row_count = 0

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
        order_id = row[COL_ORDER_ID]
        if not order_id:
            continue
        order_id = str(order_id).strip()
        row_count += 1

        if order_id not in orders:
            orders[order_id] = {'first_row': list(row), 'items': []}
        orders[order_id]['items'].append(list(row))

    print(f"Total rows: {row_count}, Unique orders: {len(orders)}")

    # Generate SQL
    lines = []
    lines.append("-- ============================================================")
    lines.append("-- Seed Data: Raw Orders (TikTok Shop Vietnam)")
    lines.append(f"-- Source: {os.path.basename(EXCEL_PATH)}")
    lines.append(f"-- Entity: {ENT_ID}")
    lines.append(f"-- Channel: {CHN_CODE}")
    lines.append(f"-- Batch: {BATCH_ID}")
    lines.append(f"-- Total Orders: {len(orders)}, Total Items: {row_count}")
    lines.append("-- Generated by generate-tiktok-order-seed.py")
    lines.append("-- ============================================================")
    lines.append("")
    lines.append("SET NAMES utf8mb4;")
    lines.append("USE db_app_sales;")
    lines.append("")

    # INSERT statements — Orders first, then Items
    lines.append("-- ---- TikTok Orders ----")

    order_lines = []
    item_lines = []
    match_stats = {'MATCHED': 0, 'UNMATCHED': 0, 'COMBO': 0}
    unmatched_skus = set()

    for order_id, data in orders.items():
        r = data['first_row']
        ord_uuid = str(uuid.uuid4())

        status, status_raw = normalize_status(r[COL_STATUS])
        order_date = sql_datetime_tiktok(r[COL_CREATED_TIME])

        # Determine completion date based on status
        completed_at = 'NULL'
        if status == 'CANCELLED':
            completed_at = sql_datetime_tiktok(r[COL_CANCELLED_TIME])
        elif status in ('COMPLETED', 'DELIVERED'):
            completed_at = sql_datetime_tiktok(r[COL_DELIVERED_TIME])

        order_lines.append(
            f"INSERT INTO drd_raw_orders "
            f"(ord_id, ent_id, chn_code, ord_channel_order_id, ord_package_id, "
            f"ord_order_date, ord_status, ord_status_raw, ord_cancel_reason, "
            f"ord_tracking_no, ord_carrier, ord_delivery_method, ord_order_type, "
            f"ord_est_delivery_date, ord_ship_date, ord_delivery_time, "
            f"ord_total_weight_kg, ord_total_vnd, "
            f"ord_shop_voucher, ord_coin_cashback, ord_shopee_voucher, "
            f"ord_promo_combo, ord_shopee_combo_discount, ord_shop_combo_discount, "
            f"ord_shopee_coin_rebate, ord_card_discount, "
            f"ord_trade_in_discount, ord_trade_in_bonus, ord_seller_trade_in_bonus, "
            f"ord_shipping_fee_est, ord_buyer_shipping_fee, ord_shopee_shipping_subsidy, "
            f"ord_return_shipping_fee, ord_total_buyer_payment, "
            f"ord_completed_at, ord_paid_at, ord_payment_method, "
            f"ord_commission_fee, ord_service_fee, ord_payment_fee, ord_deposit, "
            f"ord_province, ord_district, ord_country, "
            f"ord_import_batch_id) "
            f"VALUES ("
            f"'{ord_uuid}', '{ENT_ID}', '{CHN_CODE}', {sql_str(order_id)}, {sql_str(r[COL_PACKAGE_ID])}, "
            f"{order_date}, '{status}', {sql_str(status_raw)}, {sql_str(r[COL_CANCEL_REASON])}, "
            f"{sql_str(r[COL_TRACKING_ID])}, {sql_str(r[COL_CARRIER])}, {sql_str(r[COL_FULFILLMENT])}, {sql_str(r[COL_ORDER_TYPE])}, "
            f"NULL, {sql_datetime_tiktok(r[COL_RTS_TIME])}, {sql_datetime_tiktok(r[COL_DELIVERED_TIME])}, "
            f"NULL, {sql_decimal(r[COL_ORDER_AMOUNT])}, "
            f"NULL, NULL, NULL, "
            f"NULL, NULL, NULL, "
            f"NULL, {sql_decimal(r[COL_PAYMENT_PLATFORM_DISC])}, "
            f"NULL, NULL, NULL, "
            f"{sql_decimal(r[COL_SHIP_FEE_ORIG])}, {sql_decimal(r[COL_SHIP_FEE_AFTER])}, NULL, "
            f"NULL, {sql_decimal(r[COL_ORDER_AMOUNT])}, "
            f"{completed_at}, {sql_datetime_tiktok(r[COL_PAID_TIME])}, {sql_str(r[COL_PAYMENT_METHOD])}, "
            f"NULL, NULL, NULL, NULL, "
            f"{sql_str(r[COL_PROVINCE])}, {sql_str(r[COL_DISTRICT])}, {sql_str(r[COL_COUNTRY])}, "
            f"'{BATCH_ID}');"
        )

        # Items
        for item_row in data['items']:
            oli_uuid = str(uuid.uuid4())
            seller_sku = esc(item_row[COL_SELLER_SKU])

            # SKU matching
            sku_id = 'NULL'
            match_status = 'UNMATCHED'
            if seller_sku:
                if seller_sku.startswith('Combo') or seller_sku.startswith('COMBO') or '_GIFT_' in seller_sku.upper():
                    match_status = 'COMBO'
                    match_stats['COMBO'] += 1
                elif seller_sku in sku_map:
                    sku_id = f"'{sku_map[seller_sku]}'"
                    match_status = 'MATCHED'
                    match_stats['MATCHED'] += 1
                else:
                    match_stats['UNMATCHED'] += 1
                    unmatched_skus.add(seller_sku)
            else:
                match_stats['UNMATCHED'] += 1

            # Return status from Cancel/Return Type column
            cancel_return = esc(item_row[COL_CANCEL_RETURN_TYPE])
            return_status = cancel_return if cancel_return and cancel_return.lower() != 'cancel' else None

            item_lines.append(
                f"INSERT INTO drd_raw_order_items "
                f"(oli_id, ent_id, ord_id, sku_id, "
                f"oli_product_sku, oli_product_name, oli_variant_sku, oli_variant_name, "
                f"oli_is_bestseller, oli_weight_kg, "
                f"oli_original_price, oli_seller_discount, oli_shopee_discount, "
                f"oli_total_seller_subsidy, oli_deal_price, "
                f"oli_quantity, oli_return_quantity, oli_buyer_paid, "
                f"oli_return_status, oli_sku_match_status) "
                f"VALUES ("
                f"'{oli_uuid}', '{ENT_ID}', '{ord_uuid}', {sku_id}, "
                f"{sql_str(item_row[COL_SELLER_SKU])}, {sql_str(item_row[COL_PRODUCT_NAME])}, {sql_str(item_row[COL_SELLER_SKU])}, {sql_str(item_row[COL_VARIATION])}, "
                f"0, {sql_decimal3(item_row[COL_WEIGHT_KG])}, "
                f"{sql_decimal(item_row[COL_UNIT_PRICE])}, {sql_decimal(item_row[COL_SELLER_DISCOUNT])}, {sql_decimal(item_row[COL_PLATFORM_DISCOUNT])}, "
                f"NULL, {sql_decimal(item_row[COL_SUBTOTAL_AFTER])}, "
                f"{sql_int(item_row[COL_QUANTITY])}, {sql_int(item_row[COL_RETURN_QTY])}, {sql_decimal(item_row[COL_SUBTOTAL_AFTER])}, "
                f"{sql_str(return_status)}, '{match_status}');"
            )

    # Write orders first, then items (FK constraint order)
    lines.extend(order_lines)
    lines.append("")
    lines.append("-- ---- TikTok Order Items ----")
    lines.extend(item_lines)

    lines.append("")
    lines.append("-- EOF")

    # Write output
    with open(OUTPUT_PATH, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    total_items = sum(match_stats.values())
    print(f"\n{'='*60}")
    print(f"Generated: {OUTPUT_PATH}")
    print(f"  Orders:  {len(orders)}")
    print(f"  Items:   {total_items}")
    if total_items > 0:
        print(f"  MATCHED:   {match_stats['MATCHED']} ({match_stats['MATCHED']*100//total_items}%)")
        print(f"  UNMATCHED: {match_stats['UNMATCHED']} ({match_stats['UNMATCHED']*100//total_items}%)")
        print(f"  COMBO:     {match_stats['COMBO']} ({match_stats['COMBO']*100//total_items}%)")
    if unmatched_skus:
        print(f"\n  Unmatched SKUs ({len(unmatched_skus)}):")
        for s in sorted(unmatched_skus):
            print(f"    - {s}")
    print(f"{'='*60}")

    wb.close()

if __name__ == '__main__':
    main()
